import xlsxwriter
from openpyxl import load_workbook, Workbook
import os
from api.v1.views import app_views, needed, question_options
from api.v1.views.allquestions import DOCUMENT_FOLDER
from flask import jsonify, send_from_directory
from models.quiz import Quiz
from models.student import Student
from models import db


@app_views.route('/download-studentscore/<id>')
def get_student_score(id):
    """gets the students and their scores for a
    particular quiz"""

    quiz = db.get_or_404(Quiz, id)
    scores = quiz.scores
    filepath = DOCUMENT_FOLDER+'{}.xlsx'.format(quiz.id)
    if os.path.exists(filepath):
        workbook = load_workbook(filename=filepath)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1, value="Student's first name")
        worksheet.cell(row=1, column=2, value="Student's last name")
        worksheet.cell(row=1, column=3, value='Score')
    
    max_row = worksheet.max_row

    for n, score in enumerate(scores):
        student = db.get_or_404(Student, score.student_id)
        worksheet.cell(row=max_row + n + 1, column=1, value=student.firstname)
        worksheet.cell(row=max_row + n + 1, column=2, value=student.lastname)
        worksheet.cell(row=max_row + n + 1, column=3, value=score.score)
    
    if os.path.exists(filepath):
        workbook.save(filename=filepath)
    else:
        workbook.save(filename=filepath)

    return send_from_directory(DOCUMENT_FOLDER, '{}.xlsx'.format(quiz.id), as_attachment=True)



    